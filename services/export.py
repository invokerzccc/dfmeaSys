"""项目导出：Excel / JSON"""

import io
import json
import re
from db.database import get_db
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.formatting.rule import CellIsRule
from openpyxl.utils import get_column_letter


# ── 列定义 ──
BASIC_HEADERS   = ["序号", "L1节点", "L2节点", "L3节点", "L4节点"]
FUNC_HEADERS    = ["功能描述", "设计要求", "性能指标", "接口说明"]
DFMEA_HEADERS   = [
    "失效模式", "失效影响\n(对当前元素)", "失效影响\n(对系统/整机)",
    "严重度\nS", "特殊特性", "失效原因", "频度\nO",
    "预防控制",
    "探测控制\n设计", "探测控制\n制程", "探测控制\n验证", "探测控制\n运维",
    "探测度\nD", "RPN", "AP",
]
ACTION_HEADERS  = ["建议措施", "责任人", "期限", "措施状态", "措施效果"]
REVISED_HEADERS = ["修订S", "修订O", "修订D", "修订RPN"]
NOTES_HEADERS  = ["备注"]

ALL_HEADERS = BASIC_HEADERS + FUNC_HEADERS + DFMEA_HEADERS + ACTION_HEADERS + REVISED_HEADERS + NOTES_HEADERS

# 分组列范围 (1-indexed)
BASIC_START  = 1
BASIC_END    = len(BASIC_HEADERS)                                    # 5
FUNC_START   = BASIC_END + 1                                         # 6
FUNC_END     = BASIC_END + len(FUNC_HEADERS)                         # 9
DFMEA_START  = FUNC_END + 1                                          # 10
DFMEA_END    = FUNC_END + len(DFMEA_HEADERS)                         # 24
ACTION_START = DFMEA_END + 1                                         # 25
ACTION_END   = DFMEA_END + len(ACTION_HEADERS)                       # 29
REVISED_START = ACTION_END + 1                                       # 30
REVISED_END  = ACTION_END + len(REVISED_HEADERS)                     # 33
NOTES_START  = REVISED_END + 1                                        # 34
NOTES_END    = REVISED_END + len(NOTES_HEADERS)                       # 34
NOTES_COL_NUM = NOTES_START                                           # 34

# 关键列号 (1-indexed)
_DFMEA_BASE = FUNC_END  # 9
S_COL_NUM    = _DFMEA_BASE + 4           # 13  严重度S
O_COL_NUM    = _DFMEA_BASE + 7           # 16  频度O
D_COL_NUM    = _DFMEA_BASE + 13          # 22  探测度D
RPN_COL_NUM  = _DFMEA_BASE + 14          # 23  RPN
AP_COL_NUM   = _DFMEA_BASE + 15          # 24  AP
DET_DESIGN_COL  = _DFMEA_BASE + 9        # 18  探测控制-设计
DET_PROCESS_COL = _DFMEA_BASE + 10       # 19  探测控制-制程
DET_VERIFY_COL  = _DFMEA_BASE + 11       # 20  探测控制-验证
DET_OPS_COL     = _DFMEA_BASE + 12       # 21  探测控制-运维
REV_S_COL_NUM   = REVISED_START          # 30  修订S
REV_O_COL_NUM   = REVISED_START + 1      # 31  修订O
REV_D_COL_NUM   = REVISED_START + 2      # 32  修订D
REV_RPN_COL_NUM = REVISED_START + 3      # 33  修订RPN

S_COL_LETTER  = get_column_letter(S_COL_NUM)       # M
O_COL_LETTER  = get_column_letter(O_COL_NUM)       # P
D_COL_LETTER  = get_column_letter(D_COL_NUM)       # V
RPN_COL_LETTER = get_column_letter(RPN_COL_NUM)    # W
AP_COL_LETTER  = get_column_letter(AP_COL_NUM)     # X
REV_S_COL_LETTER = get_column_letter(REV_S_COL_NUM)
REV_O_COL_LETTER = get_column_letter(REV_O_COL_NUM)
REV_D_COL_LETTER = get_column_letter(REV_D_COL_NUM)
REV_RPN_COL_LETTER = get_column_letter(REV_RPN_COL_NUM)

# ── 颜色 ──
BASIC_FILL   = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
FUNC_FILL    = PatternFill(start_color="4A8C5C", end_color="4A8C5C", fill_type="solid")
DFMEA_FILL   = PatternFill(start_color="C0833E", end_color="C0833E", fill_type="solid")
ACTION_FILL  = PatternFill(start_color="8B6F9E", end_color="8B6F9E", fill_type="solid")
REVISED_FILL = PatternFill(start_color="8899A6", end_color="8899A6", fill_type="solid")

BASIC_GROUP_FILL   = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
FUNC_GROUP_FILL    = PatternFill(start_color="DCF0E4", end_color="DCF0E4", fill_type="solid")
DFMEA_GROUP_FILL   = PatternFill(start_color="FDE8D0", end_color="FDE8D0", fill_type="solid")
ACTION_GROUP_FILL  = PatternFill(start_color="E8DEF0", end_color="E8DEF0", fill_type="solid")
REVISED_GROUP_FILL = PatternFill(start_color="E8ECEF", end_color="E8ECEF", fill_type="solid")
NOTES_GROUP_FILL  = PatternFill(start_color="F5F5F0", end_color="F5F5F0", fill_type="solid")
NOTES_FILL        = PatternFill(start_color="8899A6", end_color="8899A6", fill_type="solid")

RPN_HIGH_FILL = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
RPN_MID_FILL  = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
AP_H_FILL = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
AP_M_FILL = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")
AP_L_FILL = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")

HEADER_FONT   = Font(size=9, bold=True, color="FFFFFF")
GROUP_FONT_S  = Font(size=10, bold=True, color="2C5282")
GROUP_FONT_F  = Font(size=10, bold=True, color="2D5A3A")
GROUP_FONT_D  = Font(size=10, bold=True, color="8B5A28")
GROUP_FONT_A  = Font(size=10, bold=True, color="5A3E7A")
GROUP_FONT_R  = Font(size=10, bold=True, color="556677")
GROUP_FONT_N  = Font(size=10, bold=True, color="888888")
DATA_FONT      = Font(size=9)
PATH_FILL      = PatternFill(start_color="F7FAFC", end_color="F7FAFC", fill_type="solid")
DETECT_FILLS   = {
    DET_DESIGN_COL:  PatternFill(start_color="F7F5F0", end_color="F7F5F0", fill_type="solid"),
    DET_PROCESS_COL: PatternFill(start_color="F7F5F0", end_color="F7F5F0", fill_type="solid"),
    DET_VERIFY_COL:  PatternFill(start_color="F7F5F0", end_color="F7F5F0", fill_type="solid"),
    DET_OPS_COL:     PatternFill(start_color="F7F5F0", end_color="F7F5F0", fill_type="solid"),
}

THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)


# ── AP 映射 ──
# 探测控制阶段 → 列号
_DETECT_STAGE_COLS = {
    "设计": DET_DESIGN_COL,
    "制程": DET_PROCESS_COL,
    "验证": DET_VERIFY_COL,
    "运维": DET_OPS_COL,
    "来料检验": DET_PROCESS_COL,
    "产线检测": DET_PROCESS_COL,
    "系统监控": DET_OPS_COL,
}


def _calc_ap(s, o, d):
    """AIAG-VDA AP 查表"""
    rules = [
        (9, 10, 9, 10, 1, 10, 'H'),
        (9, 10, 7,  8, 1, 10, 'H'),
        (9, 10, 4,  6, 1,  9, 'H'), (9, 10, 4, 6, 10, 10, 'M'),
        (9, 10, 2,  3, 1,  7, 'H'), (9, 10, 2, 3,  8,  9, 'M'), (9, 10, 2, 3, 10, 10, 'L'),
        (9, 10, 1,  1, 1,  2, 'H'), (9, 10, 1, 1,  3,  5, 'M'), (9, 10, 1, 1,  6, 10, 'L'),
        (7, 8, 9, 10, 1, 10, 'H'),
        (7, 8, 7,  8, 1,  8, 'H'), (7, 8, 7, 8,  9, 10, 'M'),
        (7, 8, 4,  6, 1,  5, 'H'), (7, 8, 4, 6,  6, 10, 'M'),
        (7, 8, 2,  3, 1,  3, 'H'), (7, 8, 2, 3,  4,  8, 'M'), (7, 8, 2, 3,  9, 10, 'L'),
        (7, 8, 1,  1, 1,  5, 'M'), (7, 8, 1, 1,  6, 10, 'L'),
        (4, 6, 9, 10, 1, 10, 'H'),
        (4, 6, 7,  8, 1,  7, 'H'), (4, 6, 7, 8,  8, 10, 'M'),
        (4, 6, 4,  6, 1,  3, 'H'), (4, 6, 4, 6,  4, 10, 'M'),
        (4, 6, 2,  3, 1,  1, 'H'), (4, 6, 2, 3,  2,  7, 'M'), (4, 6, 2, 3,  8, 10, 'L'),
        (4, 6, 1,  1, 1,  4, 'M'), (4, 6, 1, 1,  5, 10, 'L'),
        (2, 3, 9, 10, 1, 10, 'H'),
        (2, 3, 7,  8, 1,  5, 'H'), (2, 3, 7, 8,  6, 10, 'M'),
        (2, 3, 4,  6, 1,  2, 'H'), (2, 3, 4, 6,  3,  8, 'M'), (2, 3, 4, 6,  9, 10, 'L'),
        (2, 3, 2,  3, 1,  1, 'H'), (2, 3, 2, 3,  2,  5, 'M'), (2, 3, 2, 3,  6, 10, 'L'),
        (2, 3, 1,  1, 1,  2, 'M'), (2, 3, 1, 1,  3, 10, 'L'),
        (1, 1, 9, 10, 1, 10, 'H'),
        (1, 1, 7,  8, 1,  5, 'H'), (1, 1, 7, 8,  6, 10, 'M'),
        (1, 1, 4,  6, 1,  2, 'H'), (1, 1, 4, 6,  3,  8, 'M'), (1, 1, 4, 6,  9, 10, 'L'),
        (1, 1, 2,  3, 1,  1, 'H'), (1, 1, 2, 3,  2,  5, 'M'), (1, 1, 2, 3,  6, 10, 'L'),
        (1, 1, 1,  1, 1,  2, 'M'), (1, 1, 1, 1,  3, 10, 'L'),
    ]
    for s1, s2, o1, o2, d1, d2, ap in rules:
        if s1 <= s <= s2 and o1 <= o <= o2 and d1 <= d <= d2:
            return ap
    return 'L'


def _create_ap_sheet(wb):
    """创建隐藏的 AP 查表 sheet，用于 Excel 公式直接 INDEX 查询。
    数据从 row 2 开始，row = S*100 + O*10 + D - 109。
    """
    ws = wb.create_sheet("AP")
    ws.sheet_state = 'hidden'
    ws['A1'] = 'Key'
    ws['B1'] = 'AP'
    for s in range(1, 11):
        for o in range(1, 11):
            for d in range(1, 11):
                key = s * 100 + o * 10 + d
                ap = _calc_ap(s, o, d)
                row = key - 109  # key=111 → row 2
                ws.cell(row=row, column=1, value=key)
                ws.cell(row=row, column=2, value=ap)


def _build_hierarchy(conn, project_id):
    """构建 node_id → 层级信息 的映射。"""
    nodes = conn.execute(
        "SELECT id, parent_id, name FROM structure_node WHERE project_id = ? ORDER BY order_index",
        (project_id,),
    ).fetchall()

    node_map = {n["id"]: {"name": n["name"], "parent_id": n["parent_id"]} for n in nodes}

    def get_levels(node_id):
        parts = []
        cur = node_id
        while cur:
            info = node_map.get(cur)
            if not info:
                break
            parts.append(info["name"])
            cur = info["parent_id"]
        parts.reverse()
        while len(parts) < 4:
            parts.append("")
        if len(parts) > 4:
            overflow = parts[2:-1]
            parts = [parts[0], " > ".join(overflow), parts[-1]] + [""] * (4 - 3)
            parts = parts[:4]
        return parts

    result = {}
    for nid in node_map:
        result[nid] = get_levels(nid)
    return result


def _parse_detection_control(text):
    """将探测控制文本按阶段拆分为 {col_num: formatted_text}。"""
    result = {}
    if not text:
        return result
    for line in text.split('\n'):
        line = line.strip()
        if not line:
            continue
        m = re.match(r'^\[(.+?)\]\s?(.*)', line)
        if m:
            tag = m.group(1)
            content = m.group(2).strip()
            col = _DETECT_STAGE_COLS.get(tag, DET_PROCESS_COL)
            if content:
                result.setdefault(col, []).append(content)
        else:
            result.setdefault(DET_PROCESS_COL, []).append(line)
    return {c: '\n'.join('-> ' + l for l in lines) for c, lines in result.items()}


def _format_items(text):
    """给多行文本每行加 '-> ' 前缀。"""
    if not text:
        return ''
    return '\n'.join('-> ' + line.strip() for line in text.split('\n') if line.strip())


def _write_section_headers(ws, row):
    """写入分组表头（合并单元格 + 颜色）"""
    sections_config = [
        (BASIC_START, BASIC_END,   "基本信息", BASIC_GROUP_FILL,   GROUP_FONT_S),
        (FUNC_START, FUNC_END,     "功能分析", FUNC_GROUP_FILL,    GROUP_FONT_F),
        (DFMEA_START, DFMEA_END,   "DFMEA 失效分析", DFMEA_GROUP_FILL, GROUP_FONT_D),
        (ACTION_START, ACTION_END, "改进措施", ACTION_GROUP_FILL,  GROUP_FONT_A),
        (REVISED_START, REVISED_END, "修订评分", REVISED_GROUP_FILL, GROUP_FONT_R),
        (NOTES_START, NOTES_END,   "备注",     NOTES_GROUP_FILL,   GROUP_FONT_N),
    ]
    for start, end, label, fill, font in sections_config:
        if start <= end:
            start_letter = get_column_letter(start)
            end_letter = get_column_letter(end)
            if start < end:
                ws.merge_cells(f"{start_letter}{row}:{end_letter}{row}")
            cell = ws[f"{start_letter}{row}"]
            cell.value = label
            cell.font = font
            cell.fill = fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = THIN_BORDER


def _write_headers(ws, row, headers):
    """写入列标题行"""
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col_idx, value=h)
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", wrap_text=True, vertical="center")
        cell.border = THIN_BORDER
    color_mapping = [
        (BASIC_START, BASIC_END, BASIC_FILL),
        (FUNC_START, FUNC_END, FUNC_FILL),
        (DFMEA_START, DFMEA_END, DFMEA_FILL),
        (ACTION_START, ACTION_END, ACTION_FILL),
        (REVISED_START, REVISED_END, REVISED_FILL),
        (NOTES_START, NOTES_END, NOTES_FILL),
    ]
    for start, end, fill in color_mapping:
        for c in range(start, end + 1):
            cell = ws.cell(row=row, column=c)
            cell.fill = fill
            cell.font = HEADER_FONT


def _style_data_cell(cell, col_idx, value):
    """设置数据单元格样式"""
    cell.border = THIN_BORDER
    cell.font = DATA_FONT
    cell.alignment = Alignment(vertical="top", wrap_text=True)

    if BASIC_START + 1 <= col_idx <= BASIC_END:
        cell.fill = PATH_FILL

    # 探测控制 4 列底色区分
    df = DETECT_FILLS.get(col_idx)
    if df:
        cell.fill = df

    # RPN 高亮
    if col_idx == RPN_COL_NUM and isinstance(value, (int, float)):
        v = int(value)
        if v >= 200:
            cell.fill = RPN_HIGH_FILL
            cell.font = Font(size=9, bold=True, color="B91C1C")
        elif v >= 100:
            cell.fill = RPN_MID_FILL
            cell.font = Font(size=9, bold=True, color="A16207")

    # AP 高亮
    if col_idx == AP_COL_NUM:
        if value == "H":
            cell.fill = AP_H_FILL
            cell.font = Font(size=9, bold=True, color="DC2626")
        elif value == "M":
            cell.fill = AP_M_FILL
            cell.font = Font(size=9, bold=True, color="D97706")
        elif value == "L":
            cell.fill = AP_L_FILL
            cell.font = Font(size=9, bold=True, color="16A34A")

    # S/O/D 加粗居中
    if col_idx in (S_COL_NUM, O_COL_NUM, D_COL_NUM):
        cell.alignment = Alignment(horizontal="center", vertical="top")
        cell.font = Font(size=9, bold=True)


def _build_data_row(row, levels, seq):
    """构建一行数据值列表（不含公式列）"""
    detect = _parse_detection_control(row["detection_control"])

    return (
        [seq]
        + levels
        + [
            row["function_desc"],
            row["requirement"] or "",
            row["performance_spec"] or "",
            row["interface_desc"] or "",
        ]
        + [
            row["mode_desc"],
            row["local_effect"],
            row["potential_effect"],
            row["severity_S"],
            row["classification"],
            _format_items(row["potential_cause"]),
            row["occurrence_O"],
            _format_items(row["prevention_control"]),
            detect.get(DET_DESIGN_COL, ""),
            detect.get(DET_PROCESS_COL, ""),
            detect.get(DET_VERIFY_COL, ""),
            detect.get(DET_OPS_COL, ""),
            row["detection_D"],
            None,                        # RPN — formula
            None,                        # AP — formula
        ]
        + [
            _format_items(row["recommended_action"]),
            row["action_owner"],
            row["action_due_date"],
            row["action_status"],
            row["action_effect"],
        ]
        + [
            row["revised_S"] if row["revised_S"] is not None else "",
            row["revised_O"] if row["revised_O"] is not None else "",
            row["revised_D"] if row["revised_D"] is not None else "",
            "",  # 修订RPN — formula
        ]
        + [row["notes"] or ""]
    )


def _write_row_formulas(ws, rn):
    """写入 RPN / AP / 修订RPN 公式。"""
    # RPN = S × O × D
    rpn_cell = ws.cell(row=rn, column=RPN_COL_NUM)
    rpn_cell.value = f"={S_COL_LETTER}{rn}*{O_COL_LETTER}{rn}*{D_COL_LETTER}{rn}"
    rpn_cell.border = THIN_BORDER
    rpn_cell.font = DATA_FONT
    rpn_cell.alignment = Alignment(vertical="top")

    # AP = INDEX(AP!B:B, S*100+O*10+D-109)
    ap_cell = ws.cell(row=rn, column=AP_COL_NUM)
    ap_cell.value = (
        f"=INDEX(AP!B:B, {S_COL_LETTER}{rn}*100"
        f"+{O_COL_LETTER}{rn}*10+{D_COL_LETTER}{rn}-109)"
    )
    ap_cell.border = THIN_BORDER
    ap_cell.font = DATA_FONT
    ap_cell.alignment = Alignment(horizontal="center", vertical="top")

    # 修订 RPN = 修订S × 修订O × 修订D
    rev_rpn_cell = ws.cell(row=rn, column=REV_RPN_COL_NUM)
    rev_rpn_cell.value = (
        f"={REV_S_COL_LETTER}{rn}*{REV_O_COL_LETTER}{rn}"
        f"*{REV_D_COL_LETTER}{rn}"
    )
    rev_rpn_cell.border = THIN_BORDER
    rev_rpn_cell.font = DATA_FONT
    rev_rpn_cell.alignment = Alignment(vertical="top")


# ── 导出入口 ──

def export_xlsx(project_id: int):
    """导出项目 DFMEA 数据为 Excel 文件"""
    conn = get_db()
    try:
        proj = conn.execute("SELECT * FROM project WHERE id = ?", (project_id,)).fetchone()
        hierarchy = _build_hierarchy(conn, project_id)

        rows = conn.execute(
            """SELECT fm.*, fi.id AS fi_id, fi.function_desc, fi.requirement,
                      fi.performance_spec, fi.interface_desc, sn.id AS node_id
               FROM failure_mode fm
               JOIN function_item fi ON fm.function_item_id = fi.id
               JOIN structure_node sn ON fi.node_id = sn.id
               WHERE sn.project_id = ?
               ORDER BY sn.order_index, fi.order_index, fm.order_index""",
            (project_id,),
        ).fetchall()

        wb = Workbook()
        # 创建 AP 查表 sheet
        _create_ap_sheet(wb)
        ws = wb.active
        ws.title = "DFMEA"

        ncols = len(ALL_HEADERS)
        last_col_letter = get_column_letter(ncols)

        # Row 1: 标题
        ws.merge_cells(f"A1:{last_col_letter}1")
        ws["A1"] = f"DFMEA — {proj['name']}"
        ws["A1"].font = Font(size=14, bold=True)
        ws["A1"].alignment = Alignment(horizontal="center")

        # Row 2: 副标题
        ws.merge_cells(f"A2:{last_col_letter}2")
        ws["A2"] = f"导出日期: {proj['updated_at']}    失效模式总数: {len(rows)}"
        ws["A2"].font = Font(size=9, color="888888")
        ws["A2"].alignment = Alignment(horizontal="center")

        # Row 3: 分组表头
        _write_section_headers(ws, 3)

        # Row 4: 列标题
        _write_headers(ws, 4, ALL_HEADERS)

        # 数据行
        DATA_START_ROW = 5
        for i, row in enumerate(rows):
            rn = DATA_START_ROW + i
            levels = hierarchy.get(row["node_id"], ["", "", "", ""])
            vals = _build_data_row(row, levels, i + 1)

            for col_idx, v in enumerate(vals, 1):
                if v is None:
                    continue  # formula cells handled below
                cell = ws.cell(row=rn, column=col_idx, value=v if v is not None else "")
                _style_data_cell(cell, col_idx, v)

            _write_row_formulas(ws, rn)

            # RPN 着色 (基于 DB 值)
            rpn_val = row["rpn"]
            rpn_cell = ws.cell(row=rn, column=RPN_COL_NUM)
            if isinstance(rpn_val, (int, float)):
                if rpn_val >= 200:
                    rpn_cell.fill = RPN_HIGH_FILL
                    rpn_cell.font = Font(size=9, bold=True, color="B91C1C")
                elif rpn_val >= 100:
                    rpn_cell.fill = RPN_MID_FILL
                    rpn_cell.font = Font(size=9, bold=True, color="A16207")

        # AP 列条件格式 (H=红 M=黄 L=绿)
        ap_col_letter = get_column_letter(AP_COL_NUM)
        ap_range = f"{ap_col_letter}{DATA_START_ROW}:{ap_col_letter}{DATA_START_ROW + len(rows) - 1}"
        ws.conditional_formatting.add(ap_range,
            CellIsRule(operator="equal", formula=['"H"'], fill=AP_H_FILL, font=Font(size=9, bold=True, color="DC2626")))
        ws.conditional_formatting.add(ap_range,
            CellIsRule(operator="equal", formula=['"M"'], fill=AP_M_FILL, font=Font(size=9, bold=True, color="D97706")))
        ws.conditional_formatting.add(ap_range,
            CellIsRule(operator="equal", formula=['"L"'], fill=AP_L_FILL, font=Font(size=9, bold=True, color="16A34A")))

        # 列宽
        col_widths = {
            1: 5, 2: 18, 3: 18, 4: 18, 5: 18,          # 基本信息
            6: 18, 7: 20, 8: 20, 9: 16,                  # 功能分析
            10: 18, 11: 20, 12: 20,                       # 失效模式/影响
            13: 5, 14: 6, 15: 24, 16: 5,                  # S/分类/原因/O
            17: 22,                                        # 预防控制
            18: 20, 19: 20, 20: 20, 21: 20,               # 探测控制 4 列
            22: 5, 23: 7, 24: 5,                           # D/RPN/AP
            25: 20, 26: 8, 27: 10, 28: 8, 29: 16,          # 改进措施
            30: 6, 31: 6, 32: 6, 33: 8,                   # 修订评分
            34: 16,                                       # 备注
        }
        for col_idx, w in col_widths.items():
            if col_idx <= ncols:
                ws.column_dimensions[get_column_letter(col_idx)].width = w

        ws.freeze_panes = f"A{DATA_START_ROW}"
        ws.auto_filter.ref = f"A4:{last_col_letter}{DATA_START_ROW - 1 + len(rows)}"

        ws.row_dimensions[1].height = 24
        ws.row_dimensions[2].height = 18
        ws.row_dimensions[3].height = 22
        ws.row_dimensions[4].height = 32

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()

    finally:
        conn.close()


def export_template():
    """导出空 DFMEA 模板"""
    wb = Workbook()
    _create_ap_sheet(wb)
    ws = wb.active
    ws.title = "DFMEA模板"

    ncols = len(ALL_HEADERS)
    last_col_letter = get_column_letter(ncols)

    ws.merge_cells(f"A1:{last_col_letter}1")
    ws["A1"] = "DFMEA 工作表模板"
    ws["A1"].font = Font(size=14, bold=True)
    ws["A1"].alignment = Alignment(horizontal="center")

    ws.merge_cells(f"A2:{last_col_letter}2")
    ws["A2"] = "请按列填写。失效原因/预防控制/探测控制 多条目用换行分隔，每条以 -> 开头。探测控制按阶段填入对应列，无需标注阶段标签。"
    ws["A2"].font = Font(size=9, color="888888")
    ws["A2"].alignment = Alignment(horizontal="center")

    _write_section_headers(ws, 3)
    _write_headers(ws, 4, ALL_HEADERS)

    DATA_START = 5
    DATA_END = 14
    for r in range(DATA_START, DATA_END + 1):
        for c in range(1, ncols + 1):
            cell = ws.cell(row=r, column=c, value="")
            cell.border = THIN_BORDER
            cell.font = DATA_FONT
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            # 探测控制 4 列底色
            df = DETECT_FILLS.get(c)
            if df:
                cell.fill = df
        _write_row_formulas(ws, r)

    # AP 列条件格式
    ap_col_letter = get_column_letter(AP_COL_NUM)
    ap_range = f"{ap_col_letter}{DATA_START}:{ap_col_letter}{DATA_END}"
    ws.conditional_formatting.add(ap_range,
        CellIsRule(operator="equal", formula=['"H"'], fill=AP_H_FILL, font=Font(size=9, bold=True, color="DC2626")))
    ws.conditional_formatting.add(ap_range,
        CellIsRule(operator="equal", formula=['"M"'], fill=AP_M_FILL, font=Font(size=9, bold=True, color="D97706")))
    ws.conditional_formatting.add(ap_range,
        CellIsRule(operator="equal", formula=['"L"'], fill=AP_L_FILL, font=Font(size=9, bold=True, color="16A34A")))

    col_widths = [5, 18, 18, 18, 18, 18, 20, 20, 16, 18, 20, 20, 5, 6, 24, 5, 22, 20, 20, 20, 20, 5, 7, 5, 20, 8, 10, 8, 16, 6, 6, 6, 8, 16]
    for col_idx, w in enumerate(col_widths, 1):
        if col_idx <= ncols:
            ws.column_dimensions[get_column_letter(col_idx)].width = w

    ws.freeze_panes = "A5"
    ws.row_dimensions[1].height = 24
    ws.row_dimensions[2].height = 18
    ws.row_dimensions[3].height = 22
    ws.row_dimensions[4].height = 32

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


# ── JSON 导出 ──

def export_json(project_id: int):
    """导出项目完整数据为 JSON"""
    conn = get_db()
    try:
        proj = conn.execute("SELECT * FROM project WHERE id = ?", (project_id,)).fetchone()
        nodes = conn.execute("SELECT * FROM structure_node WHERE project_id = ? ORDER BY order_index", (project_id,)).fetchall()
        node_ids = [n["id"] for n in nodes]

        functions = []
        failures = []
        if node_ids:
            placeholders = ",".join("?" for _ in node_ids)
            functions = conn.execute(
                f"SELECT * FROM function_item WHERE node_id IN ({placeholders}) ORDER BY order_index",
                node_ids,
            ).fetchall()
            func_ids = [f["id"] for f in functions]
            if func_ids:
                f_placeholders = ",".join("?" for _ in func_ids)
                failures = conn.execute(
                    f"SELECT * FROM failure_mode WHERE function_item_id IN ({f_placeholders}) ORDER BY order_index",
                    func_ids,
                ).fetchall()

        refs = conn.execute("SELECT * FROM reference WHERE project_id = ?", (project_id,)).fetchall()

        # 导出关联表
        ref_nodes = []
        fm_refs = []
        if refs:
            ref_ids = [r["id"] for r in refs]
            r_placeholders = ",".join("?" for _ in ref_ids)
            ref_nodes = conn.execute(
                f"SELECT * FROM reference_node WHERE reference_id IN ({r_placeholders})",
                ref_ids,
            ).fetchall()
        if failures:
            fm_ids = [fm["id"] for fm in failures]
            f_placeholders = ",".join("?" for _ in fm_ids)
            fm_refs = conn.execute(
                f"SELECT * FROM failure_mode_reference WHERE failure_mode_id IN ({f_placeholders})",
                fm_ids,
            ).fetchall()

        return json.dumps({
            "version": "2.0",
            "exported_at": proj["updated_at"],
            "project": dict(proj),
            "structure": [dict(n) for n in nodes],
            "functions": [dict(f) for f in functions],
            "failures": [dict(fm) for fm in failures],
            "references": [dict(r) for r in refs],
            "reference_nodes": [dict(rn) for rn in ref_nodes],
            "failure_mode_references": [dict(fmr) for fmr in fm_refs],
        }, ensure_ascii=False, indent=2, default=str)

    finally:
        conn.close()
