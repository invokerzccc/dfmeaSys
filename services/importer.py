"""项目导入：JSON / Excel"""

import json
import io
from db.database import get_db
from openpyxl import load_workbook


def import_json(json_str: str, project_name: str = None):
    """从 JSON 字符串恢复项目"""
    data = json.loads(json_str)
    if "project" not in data:
        raise ValueError("无效的导出文件：缺少 project 数据")

    conn = get_db()
    try:
        # 创建项目
        p = data["project"]
        name = project_name or (p.get("name", "恢复项目") + " (恢复)")
        cur = conn.execute(
            "INSERT INTO project (name, description) VALUES (?, ?)",
            (name, p.get("description", "")),
        )
        new_project_id = cur.lastrowid

        # 恢复结构树
        node_id_map = {}
        for n in data.get("structure", []):
            old_id = n["id"]
            old_parent = n.get("parent_id")
            cur = conn.execute(
                """INSERT INTO structure_node (project_id, parent_id, name, type, part_number, description, order_index)
                   VALUES (?, NULL, ?, ?, ?, ?, ?)""",
                (new_project_id, n["name"], n["type"], n.get("part_number", ""), n.get("description", ""), n.get("order_index", 0)),
            )
            node_id_map[old_id] = cur.lastrowid

        # 更新 parent_id
        for n in data.get("structure", []):
            if n.get("parent_id") and n["parent_id"] in node_id_map:
                conn.execute(
                    "UPDATE structure_node SET parent_id = ? WHERE id = ?",
                    (node_id_map[n["parent_id"]], node_id_map[n["id"]]),
                )

        # 恢复功能项
        func_id_map = {}
        for f in data.get("functions", []):
            new_node_id = node_id_map.get(f["node_id"])
            if new_node_id is None:
                continue
            cur = conn.execute(
                """INSERT INTO function_item (node_id, function_desc, requirement, performance_spec, interface_desc, order_index)
                   VALUES (?, ?, ?, ?, ?, ?)""",
                (new_node_id, f["function_desc"], f.get("requirement", ""), f.get("performance_spec", ""),
                 f.get("interface_desc", ""), f.get("order_index", 0)),
            )
            func_id_map[f["id"]] = cur.lastrowid

        # 恢复失效模式
        fm_id_map = {}
        for fm in data.get("failures", []):
            new_fid = func_id_map.get(fm["function_item_id"])
            if new_fid is None:
                continue
            cur = conn.execute(
                """INSERT INTO failure_mode (function_item_id, mode_desc, local_effect, potential_effect,
                   severity_S, classification, potential_cause, occurrence_O,
                   prevention_control, detection_control, detection_D,
                   rpn, action_priority, recommended_action,
                   action_owner, action_due_date, action_status, action_effect,
                   revised_S, revised_O, revised_D, revised_RPN, notes, order_index)
                   VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                (new_fid, fm.get("mode_desc", ""), fm.get("local_effect", ""), fm.get("potential_effect", ""),
                 fm.get("severity_S", 1), fm.get("classification", ""), fm.get("potential_cause", ""), fm.get("occurrence_O", 1),
                 fm.get("prevention_control", ""), fm.get("detection_control", ""), fm.get("detection_D", 1),
                 fm.get("rpn", 1), fm.get("action_priority", "L"), fm.get("recommended_action", ""),
                 fm.get("action_owner", ""), fm.get("action_due_date", ""), fm.get("action_status", "未开始"), fm.get("action_effect", ""),
                 fm.get("revised_S"), fm.get("revised_O"), fm.get("revised_D"), fm.get("revised_RPN"),
                 fm.get("notes", ""), fm.get("order_index", 0)),
            )
            fm_id_map[fm["id"]] = cur.lastrowid

        # 恢复参考材料
        ref_id_map = {}
        for ref in data.get("references", []):
            # 仅保留旧格式 node_id 用于回退兼容（新版 JSON 含 reference_nodes）
            new_node_id = node_id_map.get(ref["node_id"]) if ref.get("node_id") else None
            cur = conn.execute(
                """INSERT INTO reference (project_id, node_id, title, type, file_path, url, notes)
                   VALUES (?, ?, ?, ?, ?, ?, ?)""",
                (new_project_id, new_node_id, ref["title"], ref.get("type", "其他"),
                 ref.get("file_path", ""), ref.get("url", ""), ref.get("notes", "")),
            )
            ref_id_map[ref["id"]] = cur.lastrowid

        # 恢复 reference_node 多对多关联
        for rn in data.get("reference_nodes", []):
            new_ref_id = ref_id_map.get(rn["reference_id"])
            new_node_id = node_id_map.get(rn["node_id"])
            if new_ref_id and new_node_id:
                conn.execute(
                    "INSERT OR IGNORE INTO reference_node (reference_id, node_id) VALUES (?, ?)",
                    (new_ref_id, new_node_id),
                )

        # 恢复 failure_mode_reference 多对多关联
        for fmr in data.get("failure_mode_references", []):
            new_fm_id = fm_id_map.get(fmr["failure_mode_id"])
            new_ref_id = ref_id_map.get(fmr["reference_id"])
            if new_fm_id and new_ref_id:
                conn.execute(
                    "INSERT OR IGNORE INTO failure_mode_reference (failure_mode_id, reference_id) VALUES (?, ?)",
                    (new_fm_id, new_ref_id),
                )

        conn.commit()
        return new_project_id
    finally:
        conn.close()


# ── Excel 导入 ──
# 表头关键词 → 字段名，支持换行变体（如 "严重度\nS" 或 "严重度S"）
EXCEL_HEADER_KEYWORDS = {
    # 功能
    "功能描述": "function_desc",
    "设计要求": "requirement",
    "性能指标": "performance_spec",
    "接口说明": "interface_desc",
    # 失效模式 & 影响
    "失效模式": "mode_desc",
    "对当前元素": "local_effect",
    "对系统": "potential_effect",
    "失效影响": "potential_effect",  # 通用回退
    # 评分
    "严重度": "severity_S",
    "特殊特性": "classification",
    "失效原因": "potential_cause",
    "频度": "occurrence_O",
    "预防控制": "prevention_control",
    # 探测控制 — 4 阶段列优先于通用列（关键词顺序决定匹配优先级）
    "探测控制设计": "det_design",
    "探测控制制程": "det_process",
    "探测控制验证": "det_verify",
    "探测控制运维": "det_ops",
    "探测控制": "detection_control",  # 通用回退（无阶段拆分的单列）
    "探测度": "detection_D",
    # 改进措施
    "建议措施": "recommended_action",
    "责任人": "action_owner",
    "期限": "action_due_date",
    "措施状态": "action_status",
    "措施效果": "action_effect",
    # 修订
    "修订S": "revised_S",
    "修订O": "revised_O",
    "修订D": "revised_D",
    "修订RPN": "revised_RPN",
    # 其他
    "备注": "notes",
}


def import_xlsx(file_bytes: bytes, project_id: int):
    """从 Excel 文件导入失效模式行到现有项目"""
    wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb.active

    # 找表头行 — 匹配包含关键词的列标题
    header_row = None
    headers = {}
    for row in ws.iter_rows(min_row=1, max_row=min(10, ws.max_row), values_only=False):
        for cell in row:
            sval = str(cell.value).replace("\n", "").strip() if cell.value else ""
            for kw, field in EXCEL_HEADER_KEYWORDS.items():
                if kw in sval:
                    header_row = cell.row
                    break
        if header_row:
            break

    if header_row is None:
        raise ValueError("Excel 文件中未找到 DFMEA 表头")

    # 解析表头列映射
    for cell in ws[header_row]:
        sval = str(cell.value).replace("\n", "").strip() if cell.value else ""
        # 特殊处理：失效影响有当前元素和系统两种
        if "失效影响" in sval and "当前" in sval:
            headers["local_effect"] = cell.column
        elif "失效影响" in sval and ("系统" in sval or "整机" in sval):
            headers["potential_effect"] = cell.column
        elif "探测控制" in sval and "设计" in sval and "制程" not in sval:
            headers["det_design"] = cell.column
        elif "探测控制" in sval and "制程" in sval:
            headers["det_process"] = cell.column
        elif "探测控制" in sval and "验证" in sval:
            headers["det_verify"] = cell.column
        elif "探测控制" in sval and "运维" in sval:
            headers["det_ops"] = cell.column
        else:
            for kw, field in EXCEL_HEADER_KEYWORDS.items():
                if kw in sval and field not in headers:
                    headers[field] = cell.column
                    break

    if "function_desc" not in headers:
        raise ValueError("未找到'功能描述'列")

    conn = get_db()
    try:
        # 获取项目所有节点 ID
        nodes = conn.execute("SELECT id FROM structure_node WHERE project_id = ?", (project_id,)).fetchall()
        if not nodes:
            raise ValueError("项目没有结构节点，请先创建")
        default_node_id = nodes[0]["id"]

        imported = 0
        for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
            func_desc = _get_cell(row, headers, "function_desc")
            if not func_desc:
                continue  # 空行跳过
            mode_desc = _get_cell(row, headers, "mode_desc")
            if not mode_desc:
                continue  # 没有失效模式的行跳过

            # 查找或创建功能项（含附加字段）
            func_id = _find_or_create_function(
                conn, default_node_id, func_desc,
                requirement=_get_cell(row, headers, "requirement") or "",
                performance_spec=_get_cell(row, headers, "performance_spec") or "",
                interface_desc=_get_cell(row, headers, "interface_desc") or "",
            )

            # 合并探测控制 4 阶段列 → detection_control 文本
            det_text = _combine_detection_controls(row, headers)

            # 构建失效模式数据
            S = int(_get_cell(row, headers, "severity_S") or 1)
            O = int(_get_cell(row, headers, "occurrence_O") or 1)
            D = int(_get_cell(row, headers, "detection_D") or 1)
            rpn = S * O * D
            ap = _calc_ap_from_import(S, O, D)

            conn.execute(
                """INSERT INTO failure_mode (function_item_id, mode_desc, local_effect, potential_effect,
                   severity_S, classification, potential_cause, occurrence_O,
                   prevention_control, detection_control, detection_D,
                   rpn, action_priority, recommended_action,
                   action_owner, action_due_date, action_status, action_effect,
                   revised_S, revised_O, revised_D, revised_RPN, notes)
                   VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                (func_id, mode_desc,
                 _get_cell(row, headers, "local_effect") or "",
                 _get_cell(row, headers, "potential_effect") or "",
                 S, _get_cell(row, headers, "classification") or "",
                 _get_cell(row, headers, "potential_cause") or "", O,
                 _get_cell(row, headers, "prevention_control") or "",
                 det_text, D,
                 rpn, ap, _get_cell(row, headers, "recommended_action") or "",
                 _get_cell(row, headers, "action_owner") or "",
                 _get_cell(row, headers, "action_due_date") or "",
                 _get_cell(row, headers, "action_status") or "未开始",
                 _get_cell(row, headers, "action_effect") or "",
                 int(r) if (r := _get_cell(row, headers, "revised_S")) else None,
                 int(r) if (r := _get_cell(row, headers, "revised_O")) else None,
                 int(r) if (r := _get_cell(row, headers, "revised_D")) else None,
                 int(r) if (r := _get_cell(row, headers, "revised_RPN")) else None,
                 _get_cell(row, headers, "notes") or "",
                 ),
            )
            imported += 1

        conn.commit()
        return imported
    finally:
        conn.close()


def _combine_detection_controls(row, headers):
    """将 4 阶段探测控制列合并为 detection_control 文本"""
    stages = [
        ("det_design", "[设计]"),
        ("det_process", "[制程]"),
        ("det_verify", "[验证]"),
        ("det_ops", "[运维]"),
    ]
    parts = []
    for field, tag in stages:
        val = _get_cell(row, headers, field)
        if val:
            parts.append(f"{tag} {val}")
    if parts:
        return "\n".join(parts)
    # 回退到通用单列
    return _get_cell(row, headers, "detection_control") or ""


def _get_cell(row, headers, key):
    col = headers.get(key)
    if col is None:
        return None
    idx = col - 1
    if idx >= len(row):
        return None
    val = row[idx]
    return str(val).strip() if val not in (None, "") else ""


def _find_or_create_function(conn, node_id, func_desc, requirement="", performance_spec="", interface_desc=""):
    row = conn.execute(
        "SELECT id FROM function_item WHERE node_id = ? AND function_desc = ?",
        (node_id, func_desc),
    ).fetchone()
    if row:
        # 更新已有记录的额外字段（如果导入提供了更完整的数据）
        if requirement or performance_spec or interface_desc:
            conn.execute(
                "UPDATE function_item SET requirement=?, performance_spec=?, interface_desc=? WHERE id=?",
                (requirement, performance_spec, interface_desc, row["id"]),
            )
        return row["id"]
    cur = conn.execute(
        "INSERT INTO function_item (node_id, function_desc, requirement, performance_spec, interface_desc)"
        " VALUES (?, ?, ?, ?, ?)",
        (node_id, func_desc, requirement, performance_spec, interface_desc),
    )
    return cur.lastrowid


def _calc_ap_from_import(S, O, D):
    rules = [
        (9,10, 9,10, 1,10,'H'), (9,10, 7,8, 1,10,'H'),
        (9,10, 4,6, 1,9,'H'), (9,10, 4,6, 10,10,'M'),
        (9,10, 2,3, 1,7,'H'), (9,10, 2,3, 8,9,'M'), (9,10, 2,3, 10,10,'L'),
        (9,10, 1,1, 1,2,'H'), (9,10, 1,1, 3,5,'M'), (9,10, 1,1, 6,10,'L'),
        (7,8, 9,10, 1,10,'H'), (7,8, 7,8, 1,8,'H'), (7,8, 7,8, 9,10,'M'),
        (7,8, 4,6, 1,5,'H'), (7,8, 4,6, 6,10,'M'),
        (7,8, 2,3, 1,3,'H'), (7,8, 2,3, 4,8,'M'), (7,8, 2,3, 9,10,'L'),
        (7,8, 1,1, 1,5,'M'), (7,8, 1,1, 6,10,'L'),
        (4,6, 9,10, 1,10,'H'), (4,6, 7,8, 1,7,'H'), (4,6, 7,8, 8,10,'M'),
        (4,6, 4,6, 1,3,'H'), (4,6, 4,6, 4,10,'M'),
        (4,6, 2,3, 1,1,'H'), (4,6, 2,3, 2,7,'M'), (4,6, 2,3, 8,10,'L'),
        (4,6, 1,1, 1,4,'M'), (4,6, 1,1, 5,10,'L'),
        (2,3, 9,10, 1,10,'H'), (2,3, 7,8, 1,5,'H'), (2,3, 7,8, 6,10,'M'),
        (2,3, 4,6, 1,2,'H'), (2,3, 4,6, 3,8,'M'), (2,3, 4,6, 9,10,'L'),
        (2,3, 2,3, 1,1,'H'), (2,3, 2,3, 2,5,'M'), (2,3, 2,3, 6,10,'L'),
        (2,3, 1,1, 1,2,'M'), (2,3, 1,1, 3,10,'L'),
        (1,1, 9,10, 1,10,'H'), (1,1, 7,8, 1,5,'H'), (1,1, 7,8, 6,10,'M'),
        (1,1, 4,6, 1,2,'H'), (1,1, 4,6, 3,8,'M'), (1,1, 4,6, 9,10,'L'),
        (1,1, 2,3, 1,1,'H'), (1,1, 2,3, 2,5,'M'), (1,1, 2,3, 6,10,'L'),
        (1,1, 1,1, 1,2,'M'), (1,1, 1,1, 3,10,'L'),
    ]
    for s1,s2,o1,o2,d1,d2,ap in rules:
        if s1<=S<=s2 and o1<=O<=o2 and d1<=D<=d2:
            return ap
    return 'L'
